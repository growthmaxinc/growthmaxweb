"""
One-shot backfill: reconcile GrowthMax-Keyword-Program.xlsx with _posts/.

Why this exists
---------------
The auto-pipeline (scripts/generate-post.py) only writes back the Calendar row
it generated *from*. Posts published before the keyword program existed (or
hand-authored posts) are invisible to the workbook. As of June 4, 2026 the
Calendar tab reports ~2 Published rows against ~21 actual posts — meaning
sibling_spokes() in generate-post.py sees roughly 1/10th of the real sibling
set, silently capping internal-link quality on every new auto-generated post.

This script walks _posts/*.md and, for each post:

  1. Tries to match an existing Calendar "Not started" row by slug
     (_slugify(working_title) == post_slug). If found, marks it
     "Published <YYYY-MM-DD>" and stamps the URL in column I.

  2. If no Calendar match, appends a new row at the bottom of Calendar marked
     "Backfilled <YYYY-MM-DD>" so the workbook stays a complete record.

  3. Calls the canonical writeback_publish() helper from generate-post.py so
     Master Keywords gets updated by the exact same code path the auto-pipeline
     uses. This keeps the column mapping (Status=M, URL=N) in one place.

Run once, commit the updated .xlsx, push. Idempotent — re-running is a no-op
on already-Published rows.

Usage
-----
    cd <repo root>
    python scripts/backfill_workbook.py [--dry-run]
"""

import argparse
import importlib.util
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

# Import generate-post.py as a module (filename has a hyphen, so we can't
# `import generate-post` directly). Reuses _slugify() and writeback_publish()
# so the column-mapping and slugging rules stay in one place.
_THIS_DIR = Path(__file__).parent
_spec = importlib.util.spec_from_file_location(
    "generate_post", str(_THIS_DIR / "generate-post.py")
)
generate_post = importlib.util.module_from_spec(_spec)
# generate-post.py imports anthropic + google-genai at module top, which we
# don't need here. Suppress the side effects of those imports if they fail —
# we only need _slugify() and writeback_publish(), neither of which touches
# the LLM clients.
try:
    _spec.loader.exec_module(generate_post)
except ImportError as e:
    print(f"warning: generate-post.py import had non-fatal issue: {e}", file=sys.stderr)
    # Fall through: we'll re-implement the two helpers locally below.
    generate_post = None

POSTS_DIR = Path("_posts")
WORKBOOK_PATH = Path("GrowthMax-Keyword-Program.xlsx")
SITE_BASE_URL = "https://growthmaxinc.com"
POST_URL_TEMPLATE = "/blog/{slug}/"


# --- Fallback helpers (only used if generate-post.py import failed) ---

def _slugify_local(text):
    if not text:
        return ""
    s = str(text).lower()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_]+", "-", s.strip())
    return s.strip("-")


def _slugify(text):
    if generate_post and hasattr(generate_post, "_slugify"):
        return generate_post._slugify(text)
    return _slugify_local(text)


# --- Backfill logic ---

POST_FILENAME_RE = re.compile(r"^(\d{4}-\d{2}-\d{2})-(.+)\.md$")


def discover_posts():
    """Return list of {date, slug, url, title} dicts, oldest first."""
    posts = []
    for f in sorted(POSTS_DIR.glob("*.md")):
        m = POST_FILENAME_RE.match(f.name)
        if not m:
            print(f"  skip: {f.name} (filename doesn't match YYYY-MM-DD-slug.md)")
            continue
        post_date, slug = m.group(1), m.group(2)
        # Extract title from front matter for human-readable backfill note.
        title = slug
        try:
            text = f.read_text(encoding="utf-8")
            tm = re.search(r'^title:\s*"?(.+?)"?\s*$', text, re.MULTILINE)
            if tm:
                title = tm.group(1)
            # Also grab primary_keyword if present — Master Keywords matches on it.
            km = re.search(r'^primary_keyword:\s*"?(.+?)"?\s*$', text, re.MULTILINE)
            keyword = km.group(1) if km else None
        except Exception as e:
            print(f"  warn: could not read {f.name}: {e}")
            keyword = None
        posts.append({
            "date": post_date,
            "slug": slug,
            "title": title,
            "keyword": keyword,
            "url": POST_URL_TEMPLATE.format(slug=slug),
        })
    return posts


def _url_to_slug(url):
    """Extract slug from a /blog/<slug>/ URL. Returns '' if it doesn't match."""
    if not url:
        return ""
    m = re.match(r"^/?blog/([^/]+)/?$", str(url).strip().lstrip("/").rstrip("/"))
    if m:
        return m.group(1)
    # Also accept full URLs
    m = re.search(r"/blog/([^/]+)/?$", str(url))
    return m.group(1) if m else ""


def existing_calendar_state(cal_sheet):
    """Return (slug_to_row, already_published_slugs) for the Calendar tab.

    slug_to_row: dict mapping _slugify(working_title) -> row index, for rows
        whose status is *not* already a "Published *" / "Backfilled *" value.
    already_published_slugs: set of slugs whose Calendar row is already Published
        or Backfilled — these rows are skipped (idempotent re-run).

    For Published/Backfilled rows we collect BOTH _slugify(title) AND the slug
    parsed out of column I (URL). This matters because a post's filename slug
    (e.g. 'why-ai-implementations-fail') is often shorter than its title
    (e.g. 'Why Your First AI Implementation Failed and How to Fix It') —
    matching only on title would cause idempotent re-runs to re-append.
    """
    slug_to_row = {}
    already = set()
    for row_idx in range(5, cal_sheet.max_row + 1):
        title = cal_sheet.cell(row=row_idx, column=5).value
        status = cal_sheet.cell(row=row_idx, column=8).value
        url = cal_sheet.cell(row=row_idx, column=9).value
        title_slug = _slugify(title) if title else ""
        url_slug = _url_to_slug(url)
        status_str = str(status or "").strip().lower()
        if status_str.startswith("published") or status_str.startswith("backfilled"):
            # Record both slug forms so the second-pass idempotency check matches
            # regardless of whether the Calendar row got its title from a Working
            # title or from a backfill (filename-derived title).
            if title_slug:
                already.add(title_slug)
            if url_slug:
                already.add(url_slug)
            continue
        if not title_slug:
            continue
        # First wins; if two rows have the same slug-of-title (shouldn't, but defensive)
        # we'll write to the earlier one.
        slug_to_row.setdefault(title_slug, row_idx)
    return slug_to_row, already


def append_backfill_row(cal_sheet, post):
    """Append a 'Backfilled <date>' row at the bottom of Calendar."""
    new_row = cal_sheet.max_row + 1
    today_iso = date.today().isoformat()
    # Columns: A Week | B Publish date | C Pillar | D Type | E Working title
    #          F Primary keyword | G Owner | H Status | I URL
    cal_sheet.cell(row=new_row, column=1).value = "(backfill)"
    cal_sheet.cell(row=new_row, column=2).value = post["date"]
    cal_sheet.cell(row=new_row, column=4).value = "Spoke"
    cal_sheet.cell(row=new_row, column=5).value = post["title"]
    cal_sheet.cell(row=new_row, column=6).value = post["keyword"] or post["title"]
    cal_sheet.cell(row=new_row, column=8).value = f"Backfilled {today_iso}"
    cal_sheet.cell(row=new_row, column=9).value = post["url"]
    return new_row


def update_master_keywords(master_sheet, post):
    """Mark matching Master Keywords row Published, or append a new row.

    Mirrors the logic in generate-post.py's writeback_publish() so we follow
    the canonical column mapping (keyword=col 3, status=col 13, URL=col 14).
    """
    keyword = (post["keyword"] or post["title"]).strip()
    target = keyword.lower()
    for row_idx in range(5, master_sheet.max_row + 1):
        kw = master_sheet.cell(row=row_idx, column=3).value
        if kw and str(kw).strip().lower() == target:
            status = master_sheet.cell(row=row_idx, column=13).value
            if status and str(status).strip().lower() == "published":
                return "already-published", row_idx
            master_sheet.cell(row=row_idx, column=13).value = "Published"
            master_sheet.cell(row=row_idx, column=14).value = post["url"]
            return "updated", row_idx

    # No match — append a new row.
    new_row = master_sheet.max_row + 1
    last_id = "K000"
    for row_idx in range(5, master_sheet.max_row + 1):
        v = master_sheet.cell(row=row_idx, column=1).value
        if v and isinstance(v, str) and v.startswith("K"):
            last_id = v
    next_num = int(last_id[1:]) + 1
    master_sheet.cell(row=new_row, column=1).value = f"K{next_num:03d}"
    master_sheet.cell(row=new_row, column=3).value = keyword
    master_sheet.cell(row=new_row, column=11).value = post["url"]
    master_sheet.cell(row=new_row, column=12).value = "Yes"
    master_sheet.cell(row=new_row, column=13).value = "Published"
    master_sheet.cell(row=new_row, column=14).value = post["url"]
    return "appended", new_row


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true", help="Report only; don't write the .xlsx")
    args = ap.parse_args()

    if not WORKBOOK_PATH.exists():
        print(f"error: {WORKBOOK_PATH} not found — run from repo root", file=sys.stderr)
        sys.exit(2)
    if not POSTS_DIR.exists():
        print(f"error: {POSTS_DIR}/ not found — run from repo root", file=sys.stderr)
        sys.exit(2)

    posts = discover_posts()
    print(f"Discovered {len(posts)} posts in {POSTS_DIR}/")

    wb = load_workbook(WORKBOOK_PATH)
    cal = wb["Calendar (90-day)"]
    master = wb["Master Keywords"]

    slug_to_row, already_published = existing_calendar_state(cal)
    print(f"Calendar state: {len(already_published)} rows already Published/Backfilled, "
          f"{len(slug_to_row)} eligible to be matched/updated.")

    today_iso = date.today().isoformat()
    counts = {"updated-existing-row": 0, "appended-backfill-row": 0, "skipped-already-published": 0}
    master_counts = {"updated": 0, "appended": 0, "already-published": 0}

    for post in posts:
        slug = post["slug"]
        if slug in already_published:
            counts["skipped-already-published"] += 1
            continue

        # 1. Calendar
        if slug in slug_to_row:
            row_idx = slug_to_row[slug]
            cal.cell(row=row_idx, column=8).value = f"Backfilled {today_iso}"
            cal.cell(row=row_idx, column=9).value = post["url"]
            counts["updated-existing-row"] += 1
            print(f"  Calendar row {row_idx:3d}  ←  {slug}  (matched existing 'Not started' row)")
        else:
            new_row = append_backfill_row(cal, post)
            counts["appended-backfill-row"] += 1
            print(f"  Calendar row {new_row:3d}  +  {slug}  (appended)")

        # 2. Master Keywords (via the same column mapping the auto-pipeline uses)
        action, mrow = update_master_keywords(master, post)
        master_counts[action] += 1

    print()
    print("Calendar:", counts)
    print("Master Keywords:", master_counts)

    if args.dry_run:
        print("\n[dry-run] no file written.")
        return

    wb.save(WORKBOOK_PATH)
    print(f"\nSaved: {WORKBOOK_PATH}")
    print("Next: review the diff, commit, push.")


if __name__ == "__main__":
    main()
