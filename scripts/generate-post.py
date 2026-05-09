"""
Blog post generator for GrowthMax Inc.
Generates SEO/AEO-optimized blog posts using Claude API.
Called by GitHub Actions on a schedule (Mon/Wed/Fri).

Workbook-driven: reads the next "Not started" row from the Calendar (90-day) tab
of GrowthMax-Keyword-Program.xlsx, generates a post against that keyword/pillar,
and writes back the publish status + live URL to both Calendar and Master Keywords.
"""

import os
import sys
import json
import re
from datetime import datetime, date
from pathlib import Path
from anthropic import Anthropic
from openpyxl import load_workbook

# Local module — Gemini Imagen-based hero image generator (canonical path).
# The legacy flat-vector PIL renderer at generate-hero-image.py is retained
# as a non-canonical fallback (see scripts/skills/growthmax-imagery/SKILL.md).
sys.path.insert(0, str(Path(__file__).parent))
import importlib.util
_gemini_spec = importlib.util.spec_from_file_location(
    "gemini_hero", str(Path(__file__).parent / "generate-hero-image-gemini.py")
)
gemini_hero = importlib.util.module_from_spec(_gemini_spec)
_gemini_spec.loader.exec_module(gemini_hero)

# AEO self-validation: imported from the same audit module that CI uses,
# so the pipeline enforces the exact same standard as humans + CI.
from seo_aeo_audit import validate_post_data, AEO_MIN, AEO_MAX, META_DESC_MAX

# Maximum number of regeneration attempts when AEO validation fails.
MAX_GENERATION_RETRIES = 3

# --- Configuration ---

POSTS_DIR = Path("_posts")
WORKBOOK_PATH = Path("GrowthMax-Keyword-Program.xlsx")
SITE_BASE_URL = "https://growthmaxinc.com"
POST_URL_TEMPLATE = "/blog/{slug}/"  # matches _config.yml permalink

# Pillar metadata: anchor URL, label, and the sibling spokes that already live.
# This drives the "must include a link to the pillar page" and
# "include 2 sibling links" requirements baked into the prompt.
PILLARS = {
    "1.0": {
        "name": "Enterprise AI Strategy",
        "url": "/ai-agents-for-business/",
        "category": "AI Strategy",
    },
    "2.0": {
        "name": "Custom AI Agents for Enterprise",
        "url": "/solutions/agent-development/",
        "category": "Implementation",
    },
    "3.0": {
        "name": "AI Adoption Playbook",
        "url": "/resources/ai-adoption-playbook/",
        "category": "People & Culture",
    },
    "4.0": {
        "name": "Enterprise AI Literacy & Training",
        "url": "/solutions/foundations/",
        "category": "Getting Started",
    },
    "5.0": {
        "name": "AI Partnership (Augmentation)",
        "url": "/partnership/",
        "category": "Partnership Model",
    },
}

BRAND_VOICE = """
You are writing for GrowthMax Inc, an AI consultancy that builds custom AI agents
and offers AI training bootcamps. The core brand message is "Partnership. Not Replacement."

Writing style:
- Warm, confident, direct — like a knowledgeable colleague, not a sales brochure
- Short paragraphs, clear structure
- Avoid: jargon soup, fear-mongering, over-promising, "revolutionize," "disrupt," "game-changer"
- Preferred vocabulary: augment, amplify, partner, judgment, expertise, outcomes, coherence
- Use em dashes (---) sparingly for emphasis
- Always frame AI as augmenting human expertise, never replacing it

Core messaging pillars:
1. Partnership, not replacement — AI augments human expertise
2. Clarity over hype — speak plainly about what AI can and can't do
3. Outcomes-focused — emphasize real results
4. Empathy for the human side — acknowledge AI adoption anxiety
"""


# --- Workbook helpers ---

def _pillar_key(raw):
    """Normalize pillar values from xlsx (may be float 1.0 or string '1.0')."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return f"{float(raw):.1f}"
    return str(raw).strip()


def load_program():
    """Load workbook and return the openpyxl Workbook + the relevant worksheets."""
    wb = load_workbook(WORKBOOK_PATH)
    return wb, {
        "calendar": wb["Calendar (90-day)"],
        "master": wb["Master Keywords"],
        "aeo": wb["AEO Questions"],
    }


def find_next_calendar_row(calendar_sheet):
    """Find the next 'Not started' row in the Calendar tab.

    Calendar columns (1-indexed):
        A Week | B Publish date | C Pillar | D Type | E Working title
        F Primary keyword | G Owner | H Status

    Header rows live above row 3; data starts at row 3.
    Returns (row_number, row_dict) or (None, None) if none available.
    """
    today = date.today()
    candidates = []
    for row_idx in range(3, calendar_sheet.max_row + 1):
        status = calendar_sheet.cell(row=row_idx, column=8).value
        if status and str(status).strip().lower() == "published":
            continue
        title = calendar_sheet.cell(row=row_idx, column=5).value
        keyword = calendar_sheet.cell(row=row_idx, column=6).value
        if not title or not keyword:
            continue
        publish_date = calendar_sheet.cell(row=row_idx, column=2).value
        # Convert datetime to date if needed
        if hasattr(publish_date, "date"):
            publish_date = publish_date.date()
        candidates.append((row_idx, publish_date, {
            "week": calendar_sheet.cell(row=row_idx, column=1).value,
            "publish_date": publish_date,
            "pillar": _pillar_key(calendar_sheet.cell(row=row_idx, column=3).value),
            "type": calendar_sheet.cell(row=row_idx, column=4).value,
            "title": title,
            "keyword": keyword,
            "owner": calendar_sheet.cell(row=row_idx, column=7).value,
            "status": status,
        }))

    if not candidates:
        return None, None

    # Prefer rows whose publish date is today or in the past (overdue first),
    # then earliest upcoming. Rows without a date sort last.
    def sort_key(item):
        _, pd, _ = item
        if pd is None:
            return (2, today)
        if pd <= today:
            return (0, pd)  # overdue/today, earliest first
        return (1, pd)      # upcoming, earliest first

    candidates.sort(key=sort_key)
    row_idx, _, row = candidates[0]
    return row_idx, row


def aeo_questions_for_pillar(aeo_sheet, pillar_key):
    """Return list of {q, a} dicts for the given pillar.

    AEO columns: A Pillar | B Question | C Answer angle | D Target page | E Schema type
    """
    out = []
    for row_idx in range(3, aeo_sheet.max_row + 1):
        if _pillar_key(aeo_sheet.cell(row=row_idx, column=1).value) != pillar_key:
            continue
        q = aeo_sheet.cell(row=row_idx, column=2).value
        a = aeo_sheet.cell(row=row_idx, column=3).value
        if q and a:
            out.append({"q": str(q).strip(), "a": str(a).strip()})
    return out


def sibling_spokes(master_sheet, pillar_key, exclude_keyword=None, limit=3):
    """Return up to `limit` published siblings in the same pillar, as
    [{"title": str, "url": str, "keyword": str}, ...]."""
    siblings = []
    for row_idx in range(3, master_sheet.max_row + 1):
        if _pillar_key(master_sheet.cell(row=row_idx, column=2).value) != pillar_key:
            continue
        keyword = master_sheet.cell(row=row_idx, column=3).value
        if not keyword or (exclude_keyword and str(keyword).strip().lower() == exclude_keyword.strip().lower()):
            continue
        status = master_sheet.cell(row=row_idx, column=13).value  # col M
        url = master_sheet.cell(row=row_idx, column=14).value     # col N
        if status and str(status).strip().lower() == "published" and url:
            siblings.append({
                "title": str(keyword).strip(),
                "url": str(url).strip(),
                "keyword": str(keyword).strip(),
            })
        if len(siblings) >= limit:
            break
    return siblings


def writeback_publish(wb, sheets, calendar_row_idx, primary_keyword, post_url):
    """Mark Calendar row Published, and update matching Master Keywords row's
    Status + Live URL columns. Save the workbook."""
    today_iso = date.today().isoformat()

    # Calendar Status (col H) -> "Published 2026-05-05"
    sheets["calendar"].cell(row=calendar_row_idx, column=8).value = f"Published {today_iso}"
    # Calendar also stash the URL in col I (J = column 10) so the row tells you where it landed
    sheets["calendar"].cell(row=calendar_row_idx, column=9).value = post_url

    # Match Master Keywords on keyword (case-insensitive). If no exact match,
    # add a new row at the bottom so the workbook stays a complete record.
    master = sheets["master"]
    matched = False
    target = primary_keyword.strip().lower()
    for row_idx in range(3, master.max_row + 1):
        kw = master.cell(row=row_idx, column=3).value
        if kw and str(kw).strip().lower() == target:
            master.cell(row=row_idx, column=13).value = "Published"
            master.cell(row=row_idx, column=14).value = post_url
            matched = True
            break

    if not matched:
        new_row = master.max_row + 1
        last_id = "K000"
        for row_idx in range(3, master.max_row + 1):
            v = master.cell(row=row_idx, column=1).value
            if v and isinstance(v, str) and v.startswith("K"):
                last_id = v
        next_num = int(last_id[1:]) + 1
        master.cell(row=new_row, column=1).value = f"K{next_num:03d}"
        # Pillar lookup not strictly needed — leave blank; bookkeeping pass can backfill.
        master.cell(row=new_row, column=3).value = primary_keyword.strip()
        master.cell(row=new_row, column=11).value = post_url   # Target page
        master.cell(row=new_row, column=12).value = "Yes"      # Page exists
        master.cell(row=new_row, column=13).value = "Published"
        master.cell(row=new_row, column=14).value = post_url

    wb.save(WORKBOOK_PATH)


# --- Post generation ---

def get_existing_posts():
    """Read existing post titles to avoid duplicates."""
    posts = []
    for f in POSTS_DIR.glob("*.md"):
        with open(f) as fh:
            content = fh.read()
            match = re.search(r'^title:\s*"?(.+?)"?\s*$', content, re.MULTILINE)
            if match:
                posts.append(match.group(1))
    return posts


def build_prompt(plan_row, aeo_qs, siblings, prior_errors=None):
    pillar = PILLARS.get(plan_row["pillar"], {})
    pillar_url = pillar.get("url", "")
    pillar_name = pillar.get("name", "")
    pillar_category = pillar.get("category", "AI Strategy")

    aeo_block = "\n".join(
        f'- "{q["q"]}" — Answer angle: {q["a"]}' for q in aeo_qs[:5]
    ) or "(no AEO questions on file for this pillar — invent 3 strong question-style H2s)"

    sibling_block = "\n".join(
        f'- "{s["title"]}" — {SITE_BASE_URL}{s["url"]}' for s in siblings
    ) or "(no published siblings yet — link to the pillar page twice instead)"

    existing = get_existing_posts()
    existing_list = "\n".join(f"- {t}" for t in existing[-15:])  # last 15 to keep prompt tight

    # If a prior generation attempt failed validation, surface the specific
    # errors so this attempt can correct them. This is the retry-feedback loop.
    retry_block = ""
    if prior_errors:
        retry_block = (
            "\n\nYOUR PREVIOUS ATTEMPT FAILED THE AEO AUDIT. "
            "Fix these specific issues in this attempt:\n"
            + "\n".join(f"  - {e}" for e in prior_errors)
            + "\n"
            "The 40-60 word constraint on each question H2's first paragraph is non-negotiable — "
            "count words as you write each one.\n"
        )

    return f"""
{BRAND_VOICE}{retry_block}

Generate a complete blog post for the GrowthMax Inc website, optimized for both
search engines (SEO) and answer engines / AI assistants (AEO).

THIS POST IS A PROGRAMMED SPOKE FROM THE GROWTHMAX KEYWORD PROGRAM:
- Working title: {plan_row['title']}
- Pillar: {pillar_name} ({plan_row['pillar']})
- Pillar page URL (you MUST link to this naturally in the body): {pillar_url}
- Primary keyword (use in title, intro, and at least one H2): {plan_row['keyword']}

REQUIRED INTERNAL LINKS:
1. One inline link to the pillar page using natural anchor text that includes
   the pillar name or its keyword (e.g., "[the AI Adoption Playbook]({pillar_url})").
2. Two inline links to these published sibling spokes, woven naturally into
   sentences where the topic comes up:
{sibling_block}

REQUIRED AEO QUESTIONS:
At least 3 of your H2 headings MUST be exact-question phrasings drawn from this
list (use the answer angle as your guide for what to write directly under each):
{aeo_block}

Place a 40–60 word direct answer in the FIRST PARAGRAPH under each question H2.
That paragraph is what answer engines will quote.

WORD COUNT IS CHECKED BY AN AUTOMATED AUDIT BEFORE THIS POST IS SAVED.
Any first-paragraph answer outside 40–60 words will reject this generation
and trigger a retry. Count words as you write each one. Aim for 50 words
(middle of range = safety margin). Don't pad with filler; if a topic genuinely
needs more words, put the elaboration in the SECOND paragraph of that section.

EXISTING POSTS (do not duplicate these topics):
{existing_list}

CONTENT STRUCTURE:
1. 1100–1500 words.
2. 4–6 H2 sections (## headers in markdown). At least 3 must be question-format
   from the AEO list above.
3. Use H3 subheadings (###) within sections where it adds clarity.
4. Open with a compelling hook paragraph (no heading) that includes the primary
   keyword and directly answers the working-title question.
5. Short paragraphs (2–4 sentences max).
6. Bold key concepts as **phrase**.
7. End with a forward-looking conclusion paragraph.

SCHEMA + AEO REQUIREMENTS:
8. Category: "{pillar_category}"
9. Generate 3–5 relevant tags from: ai-strategy, implementation, organizational-change,
   people-culture, adoption, employee-experience, partnership, getting-started, leadership,
   training, agents, productivity, change-management
10. Meta description under 160 chars, include primary keyword, include a clear value prop.
11. Subtitle (one compelling line).
12. URL slug (lowercase, hyphens, no special chars, include primary keyword).
13. One-sentence "tldr" — single most important takeaway, used in the highlighted
    callout box at the top of the article.
14. 3 FAQ items (q + a) — answers 1–2 sentences. These become FAQPage schema.
    At least 1 of the 3 should be one of the AEO questions listed above.

Respond in this exact JSON format (no prose around it):
{{
    "title": "The Post Title",
    "subtitle": "A compelling subtitle",
    "description": "Meta description under 160 chars with primary keyword",
    "category": "{pillar_category}",
    "tags": ["tag1", "tag2", "tag3"],
    "keywords": "comma, separated, seo, keywords",
    "slug": "the-post-slug",
    "read_time": 7,
    "tldr": "One sentence key takeaway for the callout box",
    "pillar": "{plan_row['pillar']}",
    "pillar_page": "{pillar_url}",
    "primary_keyword": "{plan_row['keyword']}",
    "faq": [
        {{"q": "An AEO question?", "a": "A direct, concise answer."}},
        {{"q": "Another related question?", "a": "Another direct answer."}},
        {{"q": "A third related question?", "a": "Third direct answer."}}
    ],
    "content": "The full markdown content of the post (H2 headers, H3 subheaders, paragraphs, bold text, the required pillar link, the required sibling links). Do NOT include the title as an H1 — the layout handles that."
}}
"""


def generate_post(client, plan_row, aeo_qs, siblings, prior_errors=None):
    prompt = build_prompt(plan_row, aeo_qs, siblings, prior_errors=prior_errors)
    response = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )
    text = response.content[0].text
    if "```json" in text:
        text = text.split("```json")[1].split("```")[0]
    elif "```" in text:
        text = text.split("```")[1].split("```")[0]
    return json.loads(text.strip())


def generate_post_with_validation(client, plan_row, aeo_qs, siblings):
    """Generate a post and validate it against AEO rules. Retry up to
    MAX_GENERATION_RETRIES times with feedback if validation fails.

    Returns the validated post_data dict, or raises RuntimeError if all
    attempts fail (workflow exits non-zero, calendar row stays Not started).
    """
    prior_errors = None
    for attempt in range(1, MAX_GENERATION_RETRIES + 1):
        print(f"\n=== Generation attempt {attempt}/{MAX_GENERATION_RETRIES} ===")
        candidate = generate_post(client, plan_row, aeo_qs, siblings, prior_errors=prior_errors)
        passed, errors = validate_post_data(candidate)
        if passed:
            print(f"  ✓ AEO validation passed on attempt {attempt}")
            return candidate
        print(f"  ✗ AEO validation failed ({len(errors)} errors):")
        for e in errors:
            print(f"    - {e}")
        prior_errors = errors

    raise RuntimeError(
        f"Generation failed AEO validation after {MAX_GENERATION_RETRIES} attempts. "
        f"Last errors: {prior_errors}. Calendar row will remain 'Not started' for next run."
    )


# Hero image generation — Gemini Imagen via the HeroImageGenerator class
# in generate-hero-image-gemini.py. Creative direction lives in
# scripts/skills/growthmax-imagery/SKILL.md and is loaded by the generator
# on construction.
#
# If GEMINI_API_KEY is not set, the auto-pipeline still runs and saves the
# post — the post just ships with the default placeholder image until the
# secret is configured. This keeps the text-generation path robust against
# image-API outages.

def generate_hero_image(client, post_data, out_dir=Path(".")):
    if not os.environ.get("GEMINI_API_KEY"):
        print(
            "GEMINI_API_KEY not set — skipping hero image generation. "
            "Post will ship with the default placeholder.",
            file=sys.stderr,
        )
        return None, None

    slug = post_data["slug"]
    out_path = out_dir / f"blog-{slug}.jpg"
    try:
        gen = gemini_hero.HeroImageGenerator(anthropic_client=client)
        path, alt, _vars = gen.render_post(
            title=post_data.get("title", ""),
            subtitle=post_data.get("subtitle", ""),
            tldr=post_data.get("tldr", ""),
            out_path=out_path,
        )
        print(f"Generated hero image: {path}")
        return path, alt
    except Exception as e:
        print(
            f"Hero image generation failed, falling back to default placeholder: {e}",
            file=sys.stderr,
        )
        return None, None


def save_post(post_data, image_path=None, image_alt=None):
    """Save the generated post as a Jekyll markdown file. Returns (path, public_url)."""
    today = datetime.now().strftime("%Y-%m-%d")
    slug = post_data["slug"]
    filename = POSTS_DIR / f"{today}-{slug}.md"

    # Build front matter manually so we control quoting/order — yaml.dump
    # would mangle nested faq objects on some configs.
    fm_lines = [
        f'title: "{post_data["title"]}"',
        f'subtitle: "{post_data["subtitle"]}"',
        f'description: "{post_data["description"]}"',
        f'category: "{post_data["category"]}"',
        f'read_time: {post_data["read_time"]}',
        f'tags: [{", ".join(post_data["tags"])}]',
        f'keywords: "{post_data["keywords"]}"',
        f'image: "/{image_path.name if image_path else "growthMAX.PNG"}"',
        f'slug: {slug}',
    ]
    if image_alt:
        fm_lines.append(f'image_alt: "{image_alt}"')
    if post_data.get("tldr"):
        fm_lines.append(f'tldr: "{post_data["tldr"]}"')
    if post_data.get("pillar"):
        fm_lines.append(f'pillar: "{post_data["pillar"]}"')
    if post_data.get("pillar_page"):
        fm_lines.append(f'pillar_page: "{post_data["pillar_page"]}"')
    if post_data.get("primary_keyword"):
        fm_lines.append(f'primary_keyword: "{post_data["primary_keyword"]}"')
    if post_data.get("faq"):
        fm_lines.append("faq:")
        for item in post_data["faq"]:
            q = item["q"].replace('"', '\\"')
            a = item["a"].replace('"', '\\"')
            fm_lines.append(f'  - q: "{q}"')
            fm_lines.append(f'    a: "{a}"')

    content = "---\n" + "\n".join(fm_lines) + "\n---\n\n" + post_data["content"] + "\n"

    with open(filename, "w") as f:
        f.write(content)

    public_url = POST_URL_TEMPLATE.format(slug=slug)
    print(f"Generated: {filename}")
    print(f"Live URL: {public_url}")
    return filename, public_url


def main():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("Error: ANTHROPIC_API_KEY not set")
        sys.exit(1)

    POSTS_DIR.mkdir(exist_ok=True)
    client = Anthropic(api_key=api_key)

    if not WORKBOOK_PATH.exists():
        print(f"Error: {WORKBOOK_PATH} not found — pipeline requires the keyword program workbook.")
        sys.exit(1)

    wb, sheets = load_program()

    # Allow workflow_dispatch to override with a specific keyword. The script
    # will look for that keyword in the Calendar tab; if not present, it falls
    # back to free-form generation against the keyword.
    topic_override = os.environ.get("TOPIC_INPUT", "").strip() or None

    plan_row = None
    calendar_row_idx = None
    if topic_override:
        # Try to find a Calendar row whose keyword OR title contains the override.
        for row_idx in range(3, sheets["calendar"].max_row + 1):
            kw = sheets["calendar"].cell(row=row_idx, column=6).value or ""
            title = sheets["calendar"].cell(row=row_idx, column=5).value or ""
            if topic_override.lower() in str(kw).lower() or topic_override.lower() in str(title).lower():
                calendar_row_idx = row_idx
                plan_row = {
                    "week": sheets["calendar"].cell(row=row_idx, column=1).value,
                    "publish_date": sheets["calendar"].cell(row=row_idx, column=2).value,
                    "pillar": _pillar_key(sheets["calendar"].cell(row=row_idx, column=3).value),
                    "type": sheets["calendar"].cell(row=row_idx, column=4).value,
                    "title": title,
                    "keyword": kw,
                    "owner": sheets["calendar"].cell(row=row_idx, column=7).value,
                    "status": sheets["calendar"].cell(row=row_idx, column=8).value,
                }
                break
        if not plan_row:
            # Free-form: build a synthetic plan row, default to Pillar 3.
            plan_row = {
                "week": "manual",
                "publish_date": date.today(),
                "pillar": "3.0",
                "type": "Spoke (manual)",
                "title": topic_override,
                "keyword": topic_override,
                "owner": "manual",
                "status": "Not started",
            }
    else:
        calendar_row_idx, plan_row = find_next_calendar_row(sheets["calendar"])
        if plan_row is None:
            print("No 'Not started' rows in Calendar — keyword program is fully published. Nothing to do.")
            sys.exit(0)

    print(f"Selected: {plan_row['title']} (pillar {plan_row['pillar']}, kw '{plan_row['keyword']}')")

    aeo_qs = aeo_questions_for_pillar(sheets["aeo"], plan_row["pillar"])
    siblings = sibling_spokes(sheets["master"], plan_row["pillar"], exclude_keyword=plan_row["keyword"])
    print(f"Found {len(aeo_qs)} AEO questions, {len(siblings)} published siblings for pillar {plan_row['pillar']}")

    try:
        post_data = generate_post_with_validation(client, plan_row, aeo_qs, siblings)
    except RuntimeError as e:
        print(f"FATAL: {e}", file=sys.stderr)
        sys.exit(1)

    image_path, image_alt = generate_hero_image(client, post_data, out_dir=Path("."))
    _, public_url = save_post(post_data, image_path=image_path, image_alt=image_alt)

    if calendar_row_idx is not None:
        writeback_publish(wb, sheets, calendar_row_idx, post_data.get("primary_keyword") or plan_row["keyword"], public_url)
        print(f"Workbook updated: Calendar row {calendar_row_idx} marked Published; Master Keywords synced.")


if __name__ == "__main__":
    main()
